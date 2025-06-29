import os
import pdfplumber
from openpyxl import load_workbook
import docx
import pandas as pd
import logging

logger = logging.getLogger(__name__)

class DocumentProcessor:
    @staticmethod
    def extract_text(file_path: str) -> str:
        text = ""
        if file_path.endswith('.pdf'):
            with pdfplumber.open(file_path) as pdf:
                text = "\n".join(page.extract_text() for page in pdf.pages if page.extract_text())
        elif file_path.endswith(('.xlsx', '.xls')):
            wb = load_workbook(filename=file_path)
            text = "\n".join(
                " ".join(str(cell) for cell in row if cell is not None)
                for sheet in wb.sheetnames
                for row in wb[sheet].iter_rows(values_only=True)
            )
        elif file_path.endswith('.docx'):
            text = "\n".join(p.text for p in docx.Document(file_path).paragraphs)
        elif file_path.endswith('.csv'):
            text = pd.read_csv(file_path).to_string()
        else:
            raise ValueError(f"Unsupported file format: {file_path}")
        logger.info(f"Extracted text (first 500 chars): {repr(text[:500])}")
        if not text or not text.strip():
            logger.warning(f"No text extracted from file: {file_path}")
            return ""
        return text